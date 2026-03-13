"""
app.py
Aplicación Flask para convertir XLSX de e-CF a XML (DGII República Dominicana).
"""

import io
import os
import zipfile
import traceback

import pandas as pd
from flask import (Flask, render_template, request,
                   send_file, jsonify, send_from_directory)
from werkzeug.utils import secure_filename

from ecf_builder import build_ecf
from version import __version__

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024   # 16 MB máximo


ALLOWED_EXTENSIONS = {'xlsx', 'xls'}


def _allowed_file(filename: str) -> bool:
    return (
        '.' in filename
        and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
    )


@app.route('/')
def index():
    return render_template('index.html', version=__version__)


@app.route('/upload', methods=['POST'])
def upload():
    """
    Recibe el XLSX, genera XMLs por fila y devuelve un ZIP.
    En caso de error parcial, incluye un archivo errores.txt en el ZIP.
    """
    if 'file' not in request.files:
        return jsonify({'error': 'No se envió ningún archivo.'}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({'error': 'Nombre de archivo vacío.'}), 400

    if not _allowed_file(file.filename):
        return jsonify({'error': 'Solo se aceptan archivos .xlsx o .xls'}), 400

    try:
        df = pd.read_excel(file.stream, dtype=str, engine='openpyxl')
    except Exception as exc:
        return jsonify({'error': f'No se pudo leer el archivo Excel: {exc}'}), 400

    if df.empty:
        return jsonify({'error': 'El archivo Excel está vacío.'}), 400

    # Normalizar cabeceras: quitar espacios y saltos de línea
    df.columns = [str(c).strip().replace('\n', '').replace('\r', '') for c in df.columns]
    df = df.where(pd.notna(df), None)

    # Mapear variantes de nombre de columna al nombre canónico
    rename_map = {}
    for col in df.columns:
        upper = col.upper()
        if upper == 'ENCF' and col != 'eNCF':
            rename_map[col] = 'eNCF'
        elif upper == 'TIPOECF' and col != 'TipoeCF':
            rename_map[col] = 'TipoeCF'
    if rename_map:
        df = df.rename(columns=rename_map)

    xmls:   list[tuple[str, str]] = []   # (contenido_xml, nombre_archivo)
    errores: list[str]            = []

    for idx, row_series in df.iterrows():
        row = row_series.to_dict()
        fila_num = idx + 2   # +2 porque pandas empieza en 0, y fila 1 es cabecera

        try:
            xml_str, filename = build_ecf(row)
            # Evitar nombres duplicados
            seen = {n for _, n in xmls}
            if filename in seen:
                base, ext = filename.rsplit('.', 1)
                filename = f'{base}_fila{fila_num}.{ext}'
            xmls.append((xml_str, filename))
        except Exception as exc:
            errores.append(
                f'Fila {fila_num}: {exc}\n{traceback.format_exc()}'
            )

    if not xmls and errores:
        return jsonify({
            'error': 'Todos los registros fallaron.',
            'detalles': '\n\n'.join(errores)
        }), 422

    # Construir ZIP en memoria
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for content, name in xmls:
            zf.writestr(name, content.encode('utf-8'))
        if errores:
            zf.writestr('errores.txt', '\n\n---\n\n'.join(errores))
    zip_buffer.seek(0)

    return send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name='ecf_generados.zip'
    )


@app.route('/template')
def download_template():
    """Descarga una plantilla XLSX de ejemplo."""
    template_path = os.path.join(app.root_path, 'static', 'plantilla_ecf.xlsx')
    if not os.path.exists(template_path):
        _generate_template(template_path)
    return send_from_directory(
        os.path.join(app.root_path, 'static'),
        'plantilla_ecf.xlsx',
        as_attachment=True
    )


def _generate_template(path: str):
    """Genera la hoja de plantilla XLSX con todas las columnas posibles."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    os.makedirs(os.path.dirname(path), exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'eCF'

    # --- Definir columnas en orden ---
    cols = []

    # Identificacion
    cols += ['TipoeCF', 'eNCF', 'FechaVencimientoSecuencia',
             'IndicadorNotaCredito', 'IndicadorEnvioDiferido',
             'IndicadorMontoGravado', 'TipoIngresos', 'TipoPago',
             'FechaLimitePago', 'TerminoPago']
    for i in range(1, 8):
        cols += [f'FormaDePago_{i}_FormaPago', f'FormaDePago_{i}_MontoPago']
    cols += ['TipoCuentaPago', 'NumeroCuentaPago', 'BancoPago',
             'FechaDesde', 'FechaHasta', 'TotalPaginas']

    # Emisor
    cols += ['RNCEmisor', 'RazonSocialEmisor', 'NombreComercial', 'Sucursal',
             'DireccionEmisor', 'Municipio', 'Provincia',
             'TelefonoEmisor_1', 'TelefonoEmisor_2', 'TelefonoEmisor_3',
             'CorreoEmisor', 'WebSite', 'ActividadEconomica',
             'CodigoVendedor', 'NumeroFacturaInterna', 'NumeroPedidoInterno',
             'ZonaVenta', 'RutaVenta', 'InformacionAdicionalEmisor', 'FechaEmision']

    # Comprador
    cols += ['RNCComprador', 'IdentificadorExtranjero', 'RazonSocialComprador',
             'ContactoComprador', 'CorreoComprador', 'DireccionComprador',
             'MunicipioComprador', 'ProvinciaComprador', 'PaisComprador',
             'FechaEntrega', 'ContactoEntrega', 'DireccionEntrega',
             'TelefonoAdicional', 'FechaOrdenCompra', 'NumeroOrdenCompra',
             'CodigoInternoComprador', 'ResponsablePago',
             'InformacionAdicionalComprador']

    # InformacionesAdicionales
    cols += ['IA_FechaEmbarque', 'IA_NumeroEmbarque', 'IA_NumeroContenedor',
             'IA_NumeroReferencia', 'IA_PesoBruto', 'IA_PesoNeto',
             'IA_UnidadPesoBruto', 'IA_UnidadPesoNeto', 'IA_CantidadBulto',
             'IA_UnidadBulto', 'IA_VolumenBulto', 'IA_UnidadVolumen',
             'IA_NombrePuertoEmbarque', 'IA_CondicionesEntrega',
             'IA_TotalFob', 'IA_Seguro', 'IA_Flete', 'IA_OtrosGastos',
             'IA_TotalCif', 'IA_RegimenAduanero',
             'IA_NombrePuertoSalida', 'IA_NombrePuertoDesembarque']

    # Transporte
    cols += ['TR_Conductor', 'TR_DocumentoTransporte', 'TR_Ficha', 'TR_Placa',
             'TR_RutaTransporte', 'TR_ZonaTransporte', 'TR_NumeroAlbaran',
             'TR_ViaTransporte', 'TR_PaisOrigen', 'TR_DireccionDestino',
             'TR_PaisDestino', 'TR_RNCIdentificacionCompaniaTransportista',
             'TR_NombreCompaniaTransportista', 'TR_NumeroViaje']

    # Totales
    cols += ['MontoGravadoTotal', 'MontoGravadoI1', 'MontoGravadoI2',
             'MontoGravadoI3', 'MontoExento', 'ITBIS1', 'ITBIS2', 'ITBIS3',
             'TotalITBIS', 'TotalITBIS1', 'TotalITBIS2', 'TotalITBIS3',
             'MontoImpuestoAdicional']
    for i in range(1, 6):
        cols += [f'ImpAd_{i}_TipoImpuesto',
                 f'ImpAd_{i}_TasaImpuestoAdicional',
                 f'ImpAd_{i}_MontoImpuestoSelectivoConsumoEspecifico',
                 f'ImpAd_{i}_MontoImpuestoSelectivoConsumoAdvalorem',
                 f'ImpAd_{i}_OtrosImpuestosAdicionales']
    cols += ['MontoTotal', 'MontoNoFacturable', 'MontoPeriodo',
             'SaldoAnterior', 'MontoAvancePago', 'ValorPagar',
             'TotalITBISRetenido', 'TotalISRRetencion',
             'TotalITBISPercepcion', 'TotalISRPercepcion']

    # OtraMoneda
    cols += ['OM_TipoMoneda', 'OM_TipoCambio',
             'OM_MontoGravadoTotalOtraMoneda',
             'OM_MontoGravado1OtraMoneda', 'OM_MontoGravado2OtraMoneda',
             'OM_MontoGravado3OtraMoneda', 'OM_MontoExentoOtraMoneda',
             'OM_TotalITBISOtraMoneda', 'OM_TotalITBIS1OtraMoneda',
             'OM_TotalITBIS2OtraMoneda', 'OM_TotalITBIS3OtraMoneda',
             'OM_MontoImpuestoAdicionalOtraMoneda', 'OM_MontoTotalOtraMoneda']
    for i in range(1, 4):
        cols += [f'OM_ImpAd_{i}_TipoImpuesto',
                 f'OM_ImpAd_{i}_TasaImpuestoAdicional',
                 f'OM_ImpAd_{i}_MontoEspecifico',
                 f'OM_ImpAd_{i}_MontoAdvalorem',
                 f'OM_ImpAd_{i}_OtrosMontos']

    # Items (3 de ejemplo)
    for n in range(1, 4):
        cols += [f'Item_{n}_NumeroLinea']
        for c in range(1, 3):
            cols += [f'Item_{n}_Cod_{c}_TipoCodigo',
                     f'Item_{n}_Cod_{c}_CodigoItem']
        cols += [f'Item_{n}_IndicadorFacturacion',
                 f'Item_{n}_Ret_IndicadorAgenteRetencionoPercepcion',
                 f'Item_{n}_Ret_MontoITBISRetenido',
                 f'Item_{n}_Ret_MontoISRRetenido',
                 f'Item_{n}_NombreItem',
                 f'Item_{n}_IndicadorBienoServicio',
                 f'Item_{n}_DescripcionItem',
                 f'Item_{n}_CantidadItem',
                 f'Item_{n}_UnidadMedida',
                 f'Item_{n}_CantidadReferencia',
                 f'Item_{n}_UnidadReferencia',
                 f'Item_{n}_GradosAlcohol',
                 f'Item_{n}_PrecioUnitarioReferencia',
                 f'Item_{n}_FechaElaboracion',
                 f'Item_{n}_FechaVencimientoItem',
                 f'Item_{n}_Mineria_PesoNetoKilogramo',
                 f'Item_{n}_Mineria_PesoNetoMineria',
                 f'Item_{n}_Mineria_TipoAfiliacion',
                 f'Item_{n}_Mineria_Liquidacion',
                 f'Item_{n}_PrecioUnitarioItem',
                 f'Item_{n}_DescuentoMonto',
                 f'Item_{n}_RecargoMonto',
                 f'Item_{n}_ImpTabla_1_TipoImpuesto',
                 f'Item_{n}_ImpTabla_2_TipoImpuesto',
                 f'Item_{n}_OM_PrecioOtraMoneda',
                 f'Item_{n}_OM_DescuentoOtraMoneda',
                 f'Item_{n}_OM_RecargoOtraMoneda',
                 f'Item_{n}_OM_MontoItemOtraMoneda',
                 f'Item_{n}_MontoItem']

    # InformacionReferencia
    cols += ['IR_NCFModificado', 'IR_RNCOtroContribuyente',
             'IR_FechaNCFModificado', 'IR_CodigoModificacion',
             'IR_RazonModificacion']

    # Firma
    cols += ['FechaHoraFirma']

    # --- Estilos ---
    header_font  = Font(bold=True, color='FFFFFF', name='Calibri', size=9)
    header_fill  = PatternFill('solid', fgColor='2D6A4F')
    center_align = Alignment(horizontal='center', vertical='center',
                              wrap_text=True)

    ws.row_dimensions[1].height = 30

    for col_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font        = header_font
        cell.fill        = header_fill
        cell.alignment   = center_align
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            len(col_name) + 2, 12
        )

    # Fila de ejemplo (tipo 31)
    example = {
        'TipoeCF': '31', 'eNCF': 'E310000000001',
        'FechaVencimientoSecuencia': '31-12-2025',
        'TipoIngresos': '01', 'TipoPago': '1',
        'RNCEmisor': '101000001', 'RazonSocialEmisor': 'EMPRESA EJEMPLO SRL',
        'DireccionEmisor': 'Calle 1 #100', 'Municipio': 'Santo Domingo',
        'Provincia': 'Distrito Nacional', 'TelefonoEmisor_1': '8095550001',
        'CorreoEmisor': 'facturacion@ejemplo.com',
        'ActividadEconomica': 'Comercio al por mayor', 'FechaEmision': '10-01-2025',
        'RNCComprador': '101000002', 'RazonSocialComprador': 'CLIENTE EJEMPLO SRL',
        'MontoGravadoTotal': '1000.00', 'ITBIS1': '18',
        'TotalITBIS': '180.00', 'TotalITBIS1': '180.00',
        'MontoTotal': '1180.00', 'ValorPagar': '1180.00',
        'Item_1_NumeroLinea': '1', 'Item_1_IndicadorFacturacion': '1',
        'Item_1_NombreItem': 'Producto A',
        'Item_1_IndicadorBienoServicio': '1',
        'Item_1_CantidadItem': '10', 'Item_1_PrecioUnitarioItem': '100.00',
        'Item_1_MontoItem': '1000.00',
    }
    for col_idx, col_name in enumerate(cols, start=1):
        if col_name in example:
            ws.cell(row=2, column=col_idx, value=example[col_name])

    wb.save(path)


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
